import numpy as np

from sklearn.metrics import balanced_accuracy_score, accuracy_score, f1_score, roc_auc_score, roc_curve, confusion_matrix
from sklearn.preprocessing import label_binarize

from collections import Counter

from openpyxl import Workbook
import os

from sympy.stats.sampling.sample_numpy import numpy


def caculate_metrics(labels, probs, preds, num_class):
    print(num_class)
    print(labels)
    print(probs)
    print(preds)

    # Balanced accuracy
    balanced_accuracy = balanced_accuracy_score(labels, preds)
    print(f'Balanced Accuracy: {balanced_accuracy:.4f}')

    # Top-1 Accuracy
    top_accuracy = accuracy_score(labels, preds)
    print(f'Top-1 Accuracy: {top_accuracy:.4f}')

    # F1-score（Macro, Micro, Weighted）
    f1_macro = f1_score(labels, preds, average='macro')
    f1_micro = f1_score(labels, preds, average='micro')
    f1_weighted = f1_score(labels, preds, average='weighted')
    print(f'Macro F1-score: {f1_macro:.4f}')
    print(f'Micro F1-score: {f1_micro:.4f}')
    print(f'Weighted F1-score: {f1_weighted:.4f}')

    if num_class == 2:
        # AUROC
        auroc = roc_auc_score(labels, probs[:, 1])
        auroc_macro = auroc
        auroc_micro = auroc
        auroc_weighted = auroc
        print(f'AUROC: {auroc:.4f}')

        fpr, tpr, _ = roc_curve(labels, probs[:, 1])
        roc_macro = dict()
        roc_weighted = dict()
        roc_micro = dict()
        roc_macro['fpr'] = fpr
        roc_macro['tpr'] = tpr
        roc_weighted['fpr'] = fpr
        roc_weighted['tpr'] = tpr
        roc_micro['fpr'] = fpr
        roc_micro['tpr'] = tpr
        print(f'FPR: {fpr}')
        print(f'TPR: {tpr}')

    else:
        labels_one_hot = label_binarize(labels, classes=np.arange(num_class))

        # Average AUROC
        auroc_macro = roc_auc_score(labels_one_hot, probs, multi_class='ovr', average='macro')
        auroc_micro = roc_auc_score(labels_one_hot, probs, multi_class='ovr', average='micro')
        auroc_weighted = roc_auc_score(labels_one_hot, probs, multi_class='ovr', average='weighted')
        print(f'Macro AUROC: {auroc_macro:.4f}')
        print(f'Micro AUROC: {auroc_micro:.4f}')
        print(f'Weighted AUROC: {auroc_weighted:.4f}')

        # ROC

        # Micro ROC
        roc_micro = dict()
        roc_micro['fpr'], roc_micro['tpr'], _ = roc_curve(labels_one_hot.ravel(), probs.ravel())
        print(f'Micro ROC fpr: {roc_micro["fpr"]}')
        print(f'Micro ROC tpr: {roc_micro["tpr"]}')

        # Macro ROC and Weighted ROC
        roc_macro = dict()
        roc_weighted = dict()
        fpr = dict()
        tpr = dict()

        for i in range(num_class):
            fpr[f"class_{i}"], tpr[f"class_{i}"], _ = roc_curve(labels_one_hot[:, i], probs[:, i])

        all_fpr = np.unique(np.concatenate([fpr[f"class_{i}"] for i in range(num_class)]))
        mean_tpr = np.zeros_like(all_fpr)
        weighted_tpr = np.zeros_like(all_fpr)
        for i in range(num_class):
            mean_tpr += np.interp(all_fpr, fpr[f"class_{i}"], tpr[f"class_{i}"])
            weighted_tpr += (np.sum(labels_one_hot[:, i]) / len(labels_one_hot)) * np.interp(all_fpr, fpr[f"class_{i}"], tpr[f"class_{i}"])
        mean_tpr /= num_class

        roc_macro['fpr'] = all_fpr
        roc_macro['tpr'] = mean_tpr
        roc_weighted['fpr'] = all_fpr
        roc_weighted['tpr'] = weighted_tpr
        print(f'Macro ROC fpr: {roc_macro["fpr"]}')
        print(f'Macro ROC tpr: {roc_macro["tpr"]}')
        print(f'Weighted ROC fpr: {roc_weighted["fpr"]}')
        print(f'Weighted ROC tpr: {roc_weighted["tpr"]}')

    cm = confusion_matrix(labels, preds)
    print("Confusion Matrix:")
    print(cm)

    return balanced_accuracy, top_accuracy, [f1_macro, f1_micro, f1_weighted], [auroc_macro, auroc_micro, auroc_weighted], [roc_macro, roc_micro, roc_weighted], cm


def save_metrics(result_save_path, balanced_accuracy, top_accuracy, F1, auroc, roc, cm):

    print("----------------------------save_metrics----------------------------")

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(["Balanced Accuracy", balanced_accuracy])
    ws1.append(["Top-1 Accuracy", top_accuracy])
    ws1.append(["Macro F1-score ", F1[0]])
    ws1.append(["Micro F1-score", F1[1]])
    ws1.append(["Weighted F1-score", F1[2]])
    ws1.append(["Macro AUROC", auroc[0]])
    ws1.append(["Micro AUROC", auroc[1]])
    ws1.append(["Weighted AUROC", auroc[2]])
    ws1.append(["Confusion Matrix", ''])
    for row in cm:
        ws1.append(row.tolist())

    ws2 = wb.create_sheet(title="Sheet2")
    ws2.append(
        ["Macro ROC fpr", "Macro ROC tpr", "Micro ROC fpr", "Micro ROC tpr", "Weighted ROC fpr", "Weighted ROC tpr"])
    for i in range(max(len(roc[0]['fpr']), len(roc[1]['fpr']), len(roc[2]['fpr']))):
        ws2.append(
            [roc[0]['fpr'][i] if i < len(roc[0]['fpr']) else '', roc[0]['tpr'][i] if i < len(roc[0]['tpr']) else '',
             roc[1]['fpr'][i] if i < len(roc[1]['fpr']) else '', roc[1]['tpr'][i] if i < len(roc[1]['tpr']) else '',
             roc[2]['fpr'][i] if i < len(roc[2]['fpr']) else '', roc[2]['tpr'][i] if i < len(roc[2]['tpr']) else ''])

    wb.save(os.path.join(result_save_path, 'test_result.xlsx'))


def caculate_retrieval_acc(indices_labels, query_label):

    top_1 = 0
    top_3 = 0
    top_5 = 0
    top_10 = 0
    MV_5 = 0
    MV_10 = 0

    if query_label == indices_labels[0]:
        top_1 = 1
    if query_label in indices_labels[0:3]:
        top_3 = 1
    if query_label in indices_labels[0:5]:
        top_5 = 1
    if query_label in indices_labels[0:10]:
        top_10 = 1

    top5_labels = indices_labels[0:5]
    most_common_top5_labels = Counter(top5_labels).most_common(1)[0][0]
    if query_label == most_common_top5_labels:
        MV_5 = 1

    top10_labels = indices_labels[0:10]
    most_common_top10_labels = Counter(top10_labels).most_common(1)[0][0]
    if query_label == most_common_top10_labels:
        MV_10 = 1

    return [top_1, top_3, top_5, top_10, MV_5, MV_10]


def save_result(img_result, acc_result, average_acc, save_dir):

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1.append(['Top1', average_acc[0]])
    ws1.append(['Top3', average_acc[1]])
    ws1.append(['Top5', average_acc[2]])
    ws1.append(['Top10', average_acc[3]])
    ws1.append(['MajorityVote5', average_acc[4]])
    ws1.append(['MajorityVote10', average_acc[5]])

    ws2 = wb.create_sheet(title="Sheet2")
    ws2.append(['query image', 'similar images'])
    for query, similar in img_result.items():
        ws2.append([query] + similar)

    ws3 = wb.create_sheet(title="Sheet3")
    ws3.append(['query image', 'top_1', 'top_3', 'top_5', 'top_10', 'MV_5', 'MV_10'])
    for query, acc in acc_result.items():
        ws3.append([query] + acc)

    wb.save(os.path.join(save_dir, 'retrieval_result.xlsx'))